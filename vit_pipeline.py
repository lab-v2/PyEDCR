import os
import torch.utils.data
import numpy as np
from sklearn.metrics import accuracy_score, f1_score
import pathlib
import typing
import random
import context_handlers
import models
import utils
import data_preprocessing
import re
import openpyxl  # Using openpyxl for Excel creation
from openpyxl.chart import (
    LineChart,
    Reference,
    Series,
)
import dropbox
from utils import TransferData


# token for Dropbox API. You can get the token from dropbox by go to dropbox:
# https://www.dropbox.com/developers/apps?_tk=pilot_lp&_ad=topbar4&_camp=myapps
# Notice that if your school account not work, use a different account
access_token = "Your token"
transferData = TransferData(access_token)

batch_size = 512
scheduler_gamma = 0.8
num_epochs = 5
ltn_num_epochs = 5
vit_model_names = [f'vit_{vit_model_name}' for vit_model_name in ['b_16']]
loss = "LTN_soft_marginal"
lrs=[3e-06]

files_path = '/content/drive/My Drive/' if utils.is_running_in_colab() else ''
combined_results_path = fr'{files_path}combined_results/'
individual_results_path = fr'{files_path}individual_results/'
cwd = pathlib.Path(__file__).parent.resolve()
scheduler_step_size = 1
betas = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]

# if fine_tune, specify it here:
training = True

# add path to model you want to evaluate and save in the excel file:
folder_path = f"/home/ngocbach/metacognitive_error_detection_and_correction_v2/model/vit_b_16_lr0.0001_{loss}_batch_size_{batch_size}_step_size_{scheduler_step_size}_scheduler_gamma_{scheduler_gamma}"
workbook_path = f"/home/ngocbach/metacognitive_error_detection_and_correction_v2/{loss}_batch_size_{batch_size}_Results.xlsx"

# set the baseline model path here:
if loss == "LTN_BCE":
    pretrained_path = "/home/ngocbach/metacognitive_error_detection_and_correction_v2/vit_b_16_BCE_lr0.0001.pth"
else:
    pretrained_path = "/home/ngocbach/metacognitive_error_detection_and_correction_v2/vit_b_16_softmarginal_1e-4.pth"


def print_num_inconsistencies(fine_labels: np.array,
                              coarse_labels: np.array,
                              prior: bool = True):
    inconsistencies = data_preprocessing.get_num_inconsistencies(fine_labels=fine_labels,
                                                                 coarse_labels=coarse_labels)

    print(f"Total {'prior' if prior else 'post'} inconsistencies "
          f"{utils.red_text(inconsistencies)}/{utils.red_text(len(fine_labels))} "
          f'({utils.red_text(round(inconsistencies / len(fine_labels) * 100, 2))}%)')


def get_and_print_metrics(fine_predictions: np.array,
                          coarse_predictions: np.array,
                          loss: str,
                          true_fine_data: np.array = data_preprocessing.true_fine_data,
                          true_coarse_data: np.array = data_preprocessing.true_coarse_data,
                          prior: bool = True,
                          combined: bool = True,
                          model_name: str = '',
                          lr: typing.Union[str, float] = ''):
    test_fine_accuracy = accuracy_score(y_true=true_fine_data,
                                        y_pred=fine_predictions)
    test_coarse_accuracy = accuracy_score(y_true=true_coarse_data,
                                          y_pred=coarse_predictions)
    test_fine_f1 = f1_score(y_true=true_fine_data,
                            y_pred=fine_predictions,
                            labels=range(len(data_preprocessing.fine_grain_classes)),
                            average='macro')
    test_coarse_f1 = f1_score(y_true=true_coarse_data,
                              y_pred=coarse_predictions,
                              labels=range(len(data_preprocessing.coarse_grain_classes)),
                              average='macro')

    prior_str = 'prior' if prior else 'post'
    combined_str = 'combined' if combined else 'individual'

    print((f'Main model name: {utils.blue_text(model_name)} ' if model_name != '' else '') +
          f'with {utils.blue_text(loss)} loss\n' +
          (f'with lr={utils.blue_text(lr)}\n' if lr != '' else '') +
          f'\nFine-grain {prior_str} {combined_str} accuracy: {utils.green_text(round(test_fine_accuracy * 100, 2))}%'
          f', fine-grain {prior_str} {combined_str} average f1: {utils.green_text(round(test_fine_f1 * 100, 2))}%'
          f'\nCoarse-grain {prior_str} {combined_str} accuracy: '
          f'{utils.green_text(round(test_coarse_accuracy * 100, 2))}%'
          f', coarse-grain {prior_str} {combined_str} average f1: '
          f'{utils.green_text(round(test_coarse_f1 * 100, 2))}%\n')

    print_num_inconsistencies(fine_labels=fine_predictions,
                              coarse_labels=coarse_predictions,
                              prior=prior)

    return test_fine_accuracy, test_coarse_accuracy


def save_test_files(fine_tuners: typing.Union[models.FineTuner, dict[str, models.FineTuner]],
                    combined: bool,
                    lrs: typing.Union[str, float, dict[str, typing.Union[str, float]]],
                    epoch: int,
                    test_fine_prediction: np.array,
                    test_coarse_prediction: np.array,
                    loss: str = 'BCE'):
    loss_str = f'{loss}_' if loss != 'BCE' else ''
    if combined:
        np.save(f"{combined_results_path}{fine_tuners}_{loss_str}test_fine_pred_lr{lrs}_e{epoch}.npy",
                test_fine_prediction)
        np.save(f"{combined_results_path}{fine_tuners}_{loss_str}test_coarse_pred_lr{lrs}_e{epoch}.npy",
                test_coarse_prediction)
    else:
        np.save(f"{individual_results_path}{fine_tuners['fine']}"
                f"_test_pred_lr{lrs['fine']}_e{epoch}_fine_individual.npy",
                test_fine_prediction)
        np.save(f"{individual_results_path}{fine_tuners['coarse']}"
                f"_test_pred_lr{lrs['coarse']}_e{epoch}_coarse_individual.npy",
                test_coarse_prediction)


def test_individual_models(fine_tuners: list[models.FineTuner],
                           loaders: dict[str, torch.utils.data.DataLoader],
                           devices: list[torch.device]) -> (list[int], list[int], float):
    test_loader = loaders[f'test']
    fine_fine_tuner, coarse_fine_tuner = fine_tuners

    device_1, device_2 = devices
    fine_fine_tuner.to(device_1)
    coarse_fine_tuner.to(device_2)

    fine_fine_tuner.eval()
    coarse_fine_tuner.eval()

    test_fine_prediction = []
    test_coarse_prediction = []

    test_fine_ground_truth = []
    test_coarse_ground_truth = []

    name_list = []

    print(f'Started testing...')

    with torch.no_grad():
        if utils.is_local():
            from tqdm import tqdm
            gen = tqdm(enumerate(test_loader), total=len(test_loader))
        else:
            gen = enumerate(test_loader)

        for i, data in gen:
            X, Y_fine_grain, names, Y_coarse_grain = data[0], data[1].to(device_1), data[2], data[3].to(device_2)

            Y_pred_fine_grain = fine_fine_tuner(X.to(device_1))
            Y_pred_coarse_grain = coarse_fine_tuner(X.to(device_2))

            predicted_fine = torch.max(Y_pred_fine_grain, 1)[1]
            predicted_coarse = torch.max(Y_pred_coarse_grain, 1)[1]

            test_fine_ground_truth += Y_fine_grain.tolist()
            test_coarse_ground_truth += Y_coarse_grain.tolist()

            test_fine_prediction += predicted_fine.tolist()
            test_coarse_prediction += predicted_coarse.tolist()

            name_list += names

    test_fine_accuracy, test_coarse_accuracy = (
        get_and_print_metrics(fine_predictions=test_fine_prediction,
                              coarse_predictions=test_coarse_prediction,
                              loss='Cross Entropy',
                              combined=False))

    return (test_fine_ground_truth, test_coarse_ground_truth, test_fine_prediction, test_coarse_prediction,
            test_fine_accuracy, test_coarse_accuracy)


def test_combined_model(fine_tuner: models.FineTuner,
                        loaders: dict[str, torch.utils.data.DataLoader],
                        loss: str,
                        device: torch.device) -> (list[int], list[int], list[int], list[int], float, float):
    test_loader = loaders['test']
    fine_tuner.to(device)
    fine_tuner.eval()

    test_fine_predictions = []
    test_coarse_predictions = []

    test_fine_ground_truths = []
    test_coarse_ground_truths = []

    print(f'Testing {fine_tuner} on {device}...')

    with torch.no_grad():
        if utils.is_local():
            from tqdm import tqdm
            gen = tqdm(enumerate(test_loader), total=len(test_loader))
        else:
            gen = enumerate(test_loader)

        for i, data in gen:
            X, Y_true_fine, Y_true_coarse = data[0].to(device), data[1].to(device), data[3].to(device)

            Y_pred = fine_tuner(X)
            Y_pred_fine = Y_pred[:, :len(data_preprocessing.fine_grain_classes)]
            Y_pred_coarse = Y_pred[:, len(data_preprocessing.fine_grain_classes):]

            predicted_fine = torch.max(Y_pred_fine, 1)[1]
            predicted_coarse = torch.max(Y_pred_coarse, 1)[1]

            test_fine_ground_truths += Y_true_fine.tolist()
            test_coarse_ground_truths += Y_true_coarse.tolist()

            test_fine_predictions += predicted_fine.tolist()
            test_coarse_predictions += predicted_coarse.tolist()

    test_fine_accuracy, test_coarse_accuracy = (
        get_and_print_metrics(fine_predictions=test_fine_predictions,
                              coarse_predictions=test_coarse_predictions,
                              loss=loss,
                              true_fine_data=test_fine_ground_truths,
                              true_coarse_data=test_coarse_ground_truths))

    return (test_fine_ground_truths, test_coarse_ground_truths, test_fine_predictions, test_coarse_predictions,
            test_fine_accuracy, test_coarse_accuracy)


def get_and_print_post_epoch_metrics(epoch: int,
                                     num_batches: int,
                                     train_fine_ground_truth: np.array,
                                     train_fine_prediction: np.array,
                                     train_coarse_ground_truth: np.array,
                                     train_coarse_prediction: np.array,
                                     num_fine_grain_classes: int,
                                     num_coarse_grain_classes: int,
                                     running_fine_loss: float = None,
                                     running_coarse_loss: float = None,
                                     running_total_loss: float = None):
    training_fine_accuracy = accuracy_score(y_true=train_fine_ground_truth, y_pred=train_fine_prediction)
    training_coarse_accuracy = accuracy_score(y_true=train_coarse_ground_truth, y_pred=train_coarse_prediction)
    training_fine_f1 = f1_score(y_true=train_fine_ground_truth, y_pred=train_fine_prediction,
                                labels=range(num_fine_grain_classes), average='macro')
    training_coarse_f1 = f1_score(y_true=train_coarse_ground_truth, y_pred=train_coarse_prediction,
                                  labels=range(num_coarse_grain_classes), average='macro')

    loss_str = (f'Training epoch total fine loss: {round(running_fine_loss / num_batches, 2)}'
                f'\ntraining epoch total coarse loss: {round(running_coarse_loss / num_batches, 2)}') \
        if running_fine_loss is not None else f'Training epoch total loss: {round(running_total_loss / num_batches, 2)}'
    print(f'\nEpoch {epoch + 1}/{num_epochs} done,\n'
          f'{loss_str}'
          f'\npost-epoch training fine accuracy: {round(training_fine_accuracy * 100, 2)}%'
          f', post-epoch fine f1: {round(training_fine_f1 * 100, 2)}%'
          f'\npost-epoch training coarse accuracy: {round(training_coarse_accuracy * 100, 2)}%'
          f', post-epoch coarse f1: {round(training_coarse_f1 * 100, 2)}%\n')

    return training_fine_accuracy, training_coarse_accuracy


def print_post_batch_metrics(batch_num: int,
                             num_batches: int,
                             batch_fine_grain_loss: float = None,
                             batch_coarse_grain_loss: float = None,
                             batch_total_loss: float = None):
    if batch_num > 0 and batch_num % 10 == 0:
        if batch_fine_grain_loss is not None:
            print(f'Completed batch num {batch_num}/{num_batches}, '
                  f'batch fine-grain loss: {round(batch_fine_grain_loss, 2)}, '
                  f'batch coarse-grain loss: {round(batch_coarse_grain_loss, 2)}')
        else:
            print(f'Completed batch num {batch_num}/{num_batches}, batch total loss: {round(batch_total_loss, 2)}')


def get_fine_tuning_batches(train_loader: torch.utils.data.DataLoader,
                            num_batches: int,
                            debug: bool):
    if utils.is_local():
        from tqdm import tqdm
        batches = tqdm(enumerate([list(train_loader)[0]] if debug else train_loader, 0),
                       total=num_batches)
    else:
        batches = enumerate(train_loader, 0)

    return batches


def fine_tune_individual_models(fine_tuners: list[models.FineTuner],
                                devices: list[torch.device],
                                loaders: dict[str, torch.utils.data.DataLoader],
                                num_fine_grain_classes: int,
                                num_coarse_grain_classes: int,
                                fine_lr: float = 1e-4,
                                coarse_lr: float = 1e-4,
                                save_files: bool = True,
                                debug: bool = False):
    fine_fine_tuner, coarse_fine_tuner = fine_tuners
    device_1, device_2 = devices
    fine_fine_tuner.to(device_1)
    fine_fine_tuner.train()

    coarse_fine_tuner.to(device_2)
    coarse_fine_tuner.train()

    train_loader = loaders['train']
    num_batches = len(train_loader)
    criterion = torch.nn.CrossEntropyLoss()

    fine_optimizer = torch.optim.Adam(params=fine_fine_tuner.parameters(),
                                      lr=fine_lr)
    coarse_optimizer = torch.optim.Adam(params=coarse_fine_tuner.parameters(),
                                        lr=coarse_lr)

    fine_scheduler = torch.optim.lr_scheduler.StepLR(optimizer=fine_optimizer,
                                                     step_size=scheduler_step_size,
                                                     gamma=scheduler_gamma)
    coarse_scheduler = torch.optim.lr_scheduler.StepLR(optimizer=coarse_optimizer,
                                                       step_size=scheduler_step_size,
                                                       gamma=scheduler_gamma)

    train_fine_losses = []
    train_fine_accuracies = []

    test_fine_ground_truths = []
    test_fine_accuracies = []

    train_coarse_losses = []
    train_coarse_accuracies = []

    test_coarse_ground_truths = []
    test_coarse_accuracies = []

    print(f'Started fine-tuning individual models with fine_lr={fine_lr} and coarse_lr={coarse_lr}'
          f'for {num_epochs} epochs on {device_1} and {device_2}...')

    for epoch in range(num_epochs):
        with context_handlers.TimeWrapper():
            running_fine_loss = 0.0
            running_coarse_loss = 0.0

            train_fine_predictions = []
            train_coarse_predictions = []

            train_fine_ground_truths = []
            train_coarse_ground_truths = []

            batches = get_fine_tuning_batches(train_loader=train_loader,
                                              num_batches=num_batches,
                                              debug=debug)

            for batch_num, batch in batches:
                with context_handlers.ClearCache(device=device_1):
                    with context_handlers.ClearCache(device=device_2):
                        X, Y_true_fine, Y_true_coarse = batch[0], batch[1].to(device_1), batch[3].to(device_2)

                        fine_optimizer.zero_grad()
                        coarse_optimizer.zero_grad()

                        Y_pred_fine = fine_fine_tuner(X.to(device_1))
                        Y_pred_coarse = coarse_fine_tuner(X.to(device_2))

                        batch_fine_grain_loss = criterion(Y_pred_fine, Y_true_fine)
                        batch_coarse_grain_loss = criterion(Y_pred_coarse, Y_true_coarse)

                        batch_fine_grain_loss.backward()
                        batch_coarse_grain_loss.backward()

                        fine_optimizer.step()
                        coarse_optimizer.step()

                        running_fine_loss += batch_fine_grain_loss.item()
                        running_coarse_loss += batch_coarse_grain_loss.item()

                        predicted_fine = torch.max(Y_pred_fine, 1)[1]
                        predicted_coarse = torch.max(Y_pred_coarse, 1)[1]

                        train_fine_ground_truths += Y_true_fine.tolist()
                        train_coarse_ground_truths += Y_true_coarse.tolist()

                        train_fine_predictions += predicted_fine.tolist()
                        train_coarse_predictions += predicted_coarse.tolist()

                        del X, Y_true_fine, Y_true_coarse

                        print_post_batch_metrics(batch_num=batch_num,
                                                 num_batches=num_batches,
                                                 batch_fine_grain_loss=batch_fine_grain_loss.item(),
                                                 batch_coarse_grain_loss=batch_coarse_grain_loss.item())

            true_fine_labels = np.array(train_fine_ground_truths)
            true_coarse_labels = np.array(train_coarse_ground_truths)

            predicted_fine_labels = np.array(train_fine_predictions)
            predicted_coarse_labels = np.array(train_coarse_predictions)

            training_fine_accuracy, training_coarse_accuracy = (
                get_and_print_post_epoch_metrics(epoch=epoch,
                                                 running_fine_loss=running_fine_loss,
                                                 running_coarse_loss=running_coarse_loss,
                                                 num_batches=num_batches,
                                                 train_fine_ground_truth=true_fine_labels,
                                                 train_fine_prediction=predicted_fine_labels,
                                                 train_coarse_ground_truth=true_coarse_labels,
                                                 train_coarse_prediction=predicted_coarse_labels,
                                                 num_fine_grain_classes=num_fine_grain_classes,
                                                 num_coarse_grain_classes=num_coarse_grain_classes))

            train_fine_accuracies += [training_fine_accuracy]
            train_coarse_accuracies += [training_coarse_accuracy]

            train_fine_losses += [running_fine_loss / num_batches]
            train_coarse_losses += [running_coarse_loss / num_batches]

            fine_scheduler.step()
            coarse_scheduler.step()

            (test_fine_ground_truths, test_coarse_ground_truths, test_fine_predictions, test_coarse_predictions,
             test_fine_accuracy, test_coarse_accuracy) = (
                test_individual_models(fine_tuners=fine_tuners,
                                       loaders=loaders,
                                       devices=devices))
            test_fine_accuracies += [test_fine_accuracy]
            test_coarse_accuracies += [test_coarse_accuracy]
            print('#' * 100)

            np.save(f"{individual_results_path}{fine_fine_tuner}"
                    f"_test_pred_lr{fine_lr}_e{epoch}_fine_individual.npy",
                    test_fine_predictions)
            np.save(f"{individual_results_path}{coarse_fine_tuner}"
                    f"_test_pred_lr{coarse_lr}_e{epoch}_coarse_individual.npy",
                    test_coarse_predictions)

            if save_files:
                save_test_files(fine_tuners={'fine': fine_fine_tuner,
                                             'coarse': coarse_fine_tuner},
                                combined=False,
                                lrs={'fine': fine_lr,
                                     'coarse': coarse_lr},
                                epoch=epoch,
                                test_fine_prediction=test_fine_predictions,
                                test_coarse_prediction=test_coarse_predictions)

    torch.save(fine_fine_tuner.state_dict(), f"{fine_fine_tuner}_lr{fine_lr}_fine_individual.pth")
    torch.save(coarse_fine_tuner.state_dict(), f"{coarse_fine_tuner}_lr{coarse_lr}_coarse_individual.pth")

    if not os.path.exists(f"{individual_results_path}test_true_fine_individual.npy"):
        np.save(f"{individual_results_path}test_true_fine_individual.npy", test_fine_ground_truths)
    if not os.path.exists(f"{individual_results_path}test_true_coarse_individual.npy"):
        np.save(f"{individual_results_path}test_true_coarse_individual.npy", test_coarse_ground_truths)


def fine_tune_combined_model(lrs: list[typing.Union[str, float]],
                             fine_tuner: models.FineTuner,
                             device: torch.device,
                             loaders: dict[str, torch.utils.data.DataLoader],
                             num_fine_grain_classes: int,
                             num_coarse_grain_classes: int,
                             loss: str,
                             ltn_num_epochs: int = None,
                             beta: float = 0.1,
                             save_files: bool = True,
                             debug: bool = False):
    fine_tuner.to(device)
    fine_tuner.train()
    train_loader = loaders['train']
    num_batches = len(train_loader)

    for lr in lrs:
        optimizer = torch.optim.Adam(params=fine_tuner.parameters(),
                                     lr=lr)

        scheduler = torch.optim.lr_scheduler.StepLR(optimizer=optimizer,
                                                    step_size=scheduler_step_size,
                                                    gamma=scheduler_gamma)

        alpha = num_fine_grain_classes / (num_fine_grain_classes + num_coarse_grain_classes)

        train_total_losses = []
        train_fine_losses = []
        train_coarse_losses = []

        train_fine_accuracies = []
        train_coarse_accuracies = []

        test_fine_ground_truths = []
        test_coarse_ground_truths = []

        test_fine_accuracies = []
        test_coarse_accuracies = []

        if loss.split('_')[0] == 'LTN':
            import ltn
            import ltn_support

            epochs = ltn_num_epochs
            logits_to_predicate = ltn.Predicate(ltn_support.LogitsToPredicate()).to(ltn.device)
        else:
            epochs = num_epochs

        print(f'Fine-tuning {fine_tuner} with {len(fine_tuner)} parameters for {epochs} epochs '
              f'using lr={lr} and beta={beta} and step_size={scheduler_step_size} and gamma={scheduler_gamma} on {device}...')
        print('#' * 100 + '\n')

        for epoch in range(epochs):
            with context_handlers.TimeWrapper():
                total_running_loss = torch.Tensor([0.0]).to(device)
                running_fine_loss = torch.Tensor([0.0]).to(device)
                running_coarse_loss = torch.Tensor([0.0]).to(device)

                train_fine_predictions = []
                train_coarse_predictions = []

                train_fine_ground_truths = []
                train_coarse_ground_truths = []

                batches = get_fine_tuning_batches(train_loader=train_loader,
                                                  num_batches=num_batches,
                                                  debug=debug)

                for batch_num, batch in batches:
                    with context_handlers.ClearCache(device=device):
                        X, Y_fine_grain, Y_coarse_grain = batch[0].to(device), batch[1].to(device), batch[3].to(device)
                        Y_fine_grain_one_hot = torch.nn.functional.one_hot(Y_fine_grain, num_classes=len(
                            data_preprocessing.fine_grain_classes))
                        Y_coarse_grain_one_hot = torch.nn.functional.one_hot(Y_coarse_grain, num_classes=len(
                            data_preprocessing.coarse_grain_classes))

                        Y_combine = torch.cat(tensors=[Y_fine_grain_one_hot, Y_coarse_grain_one_hot], dim=1).float()
                        optimizer.zero_grad()

                        Y_pred = fine_tuner(X)
                        Y_pred_fine_grain = Y_pred[:, :num_fine_grain_classes]
                        Y_pred_coarse_grain = Y_pred[:, num_fine_grain_classes:]

                        if loss == "weighted":
                            criterion = torch.nn.CrossEntropyLoss()

                            batch_fine_grain_loss = criterion(Y_pred_fine_grain, Y_fine_grain)
                            batch_coarse_grain_loss = criterion(Y_pred_coarse_grain, Y_coarse_grain)

                            running_fine_loss += batch_fine_grain_loss
                            running_coarse_loss += batch_coarse_grain_loss

                            batch_total_loss = alpha * batch_fine_grain_loss + (1 - alpha) * batch_coarse_grain_loss

                        elif loss == "BCE":
                            criterion = torch.nn.BCEWithLogitsLoss()
                            batch_total_loss = criterion(Y_pred, Y_combine)

                        elif loss == "CE":
                            criterion = torch.nn.CrossEntropyLoss()
                            batch_total_loss = criterion(Y_pred, Y_combine)

                        elif loss == "soft_marginal":
                            criterion = torch.nn.MultiLabelSoftMarginLoss()

                            batch_total_loss = criterion(Y_pred, Y_combine)

                        elif loss.split('_')[0] == 'LTN':
                            if loss == 'LTN_BCE':
                                criterion = torch.nn.BCEWithLogitsLoss()

                                sat_agg = ltn_support.compute_sat_normally(logits_to_predicate,
                                                                           Y_pred, Y_coarse_grain, Y_fine_grain)
                                batch_total_loss = beta * (1. - sat_agg) + (1 - beta) * (criterion(Y_pred, Y_combine))

                            if loss == "LTN_soft_marginal":
                                criterion = torch.nn.MultiLabelSoftMarginLoss()

                                sat_agg = ltn_support.compute_sat_normally(logits_to_predicate,
                                                                           Y_pred, Y_coarse_grain, Y_fine_grain)
                                batch_total_loss = beta * (1. - sat_agg) + (1 - beta) * (criterion(Y_pred, Y_combine))

                        print_post_batch_metrics(batch_num=batch_num,
                                                 num_batches=num_batches,
                                                 batch_total_loss=batch_total_loss.item())

                        batch_total_loss.backward()
                        optimizer.step()

                        total_running_loss += batch_total_loss.item()

                        predicted_fine = torch.max(Y_pred_fine_grain, 1)[1]
                        predicted_coarse = torch.max(Y_pred_coarse_grain, 1)[1]

                        train_fine_predictions += predicted_fine.tolist()
                        train_coarse_predictions += predicted_coarse.tolist()

                        train_fine_ground_truths += Y_fine_grain.tolist()
                        train_coarse_ground_truths += Y_coarse_grain.tolist()

                        del X, Y_fine_grain, Y_coarse_grain, Y_pred, Y_pred_fine_grain, Y_pred_coarse_grain

                training_fine_accuracy, training_coarse_accuracy = (
                    get_and_print_post_epoch_metrics(epoch=epoch,
                                                     running_fine_loss=running_fine_loss.item(),
                                                     running_coarse_loss=running_coarse_loss.item(),
                                                     num_batches=num_batches,
                                                     train_fine_ground_truth=np.array(train_fine_ground_truths),
                                                     train_fine_prediction=np.array(train_fine_predictions),
                                                     train_coarse_ground_truth=np.array(train_coarse_ground_truths),
                                                     train_coarse_prediction=np.array(train_coarse_predictions),
                                                     num_fine_grain_classes=num_fine_grain_classes,
                                                     num_coarse_grain_classes=num_coarse_grain_classes))

                train_fine_accuracies += [training_fine_accuracy]
                train_coarse_accuracies += [training_coarse_accuracy]

                train_total_losses += [total_running_loss.item() / num_batches]
                train_fine_losses += [running_fine_loss.item() / num_batches]
                train_coarse_losses += [running_coarse_loss.item() / num_batches]

                scheduler.step()
                (test_fine_ground_truths, test_coarse_ground_truths, test_fine_predictions, test_coarse_predictions,
                 test_fine_accuracy, test_coarse_accuracy) = (
                    test_combined_model(fine_tuner=fine_tuner,
                                        loaders=loaders,
                                        loss=loss,
                                        device=device))

                test_fine_accuracies += [test_fine_accuracy]
                test_coarse_accuracies += [test_coarse_accuracy]
                print('#' * 100)

                if (epoch == num_epochs - 1) and save_files:
                    save_test_files(fine_tuners=fine_tuner,
                                    combined=True,
                                    lrs=lr,
                                    epoch=epoch,
                                    test_fine_prediction=test_fine_predictions,
                                    test_coarse_prediction=test_coarse_predictions,
                                    loss=loss)
                    
        # Get the current working directory as the base path
        base_path = os.getcwd()

        # Create folder with configuration details if it doesn't exist
        config_folder_name = f"{fine_tuner}_lr{lr}_{loss}_batch_size_{batch_size}_step_size_{scheduler_step_size}_scheduler_gamma_{scheduler_gamma}"
        config_folder_path = os.path.join(base_path, "model", config_folder_name)
        os.makedirs(config_folder_path, exist_ok=True)


        if not os.path.exists(f"{combined_results_path}test_fine_true.npy"):
            np.save(f"{combined_results_path}test_fine_true.npy", test_fine_ground_truths)
        if not os.path.exists(f"{combined_results_path}test_coarse_true.npy"):
            np.save(f"{combined_results_path}test_coarse_true.npy", test_coarse_ground_truths)

        if loss.split('_')[0] == 'LTN':
            # Saving file in the current directory
            file_name = f"{fine_tuner}_lr{lr}_{loss}_beta_{beta}_batch_size_{batch_size}_step_size_{scheduler_step_size}_scheduler_gamma_{scheduler_gamma}"
            save_model_path = os.path.join(config_folder_path, file_name)
            torch.save(fine_tuner.state_dict(), f"{save_model_path}.pth")
            np.save(f"{save_model_path}_fine_pred.npy", test_fine_predictions)
            np.save(f"{save_model_path}_coarse_pred.npy", test_coarse_predictions)

            # transfer file
            file_based_from = save_model_path

            # from folder structure at dropbox
            file_to_folder_loss = "BCE" if loss == "LTN_BCE" else "Softmarginal"
            file_to_folder_lr = "1e-4" if lrs[0] == 1e-04 else "3e-6"
            file_based_to = f"/EDCR/combined/{file_to_folder_loss}/vit_b_16/lr_{file_to_folder_lr}/batch_size_{batch_size}/scheduler_gamma_{scheduler_gamma}/step_size_{scheduler_step_size}/beta_{beta}/" 

            # transfer model
            file_model_from = f"{file_based_from}.pth"
            file_model_to = f"{file_based_to}{file_name}.pth"
            transferData.upload_file(file_model_from, file_model_to)

            # transfer prediction
            file_fine_pred_from = f"{file_based_from}_fine_pred.npy"
            file_fine_pred_to = f"{file_based_to}{file_name}_fine_pred.npy"
            transferData.upload_file(file_fine_pred_from, file_fine_pred_to)

            file_coarse_pred_from = f"{file_based_from}_coarse_pred.npy"
            file_coarse_pred_to = f"{file_based_to}{file_name}_coarse_pred.npy"
            transferData.upload_file(file_coarse_pred_from, file_coarse_pred_to)

            print(f"predict fine and coarse and model is sent to: {file_based_to} with file name: {file_name}")
            
        else:
            torch.save(fine_tuner.state_dict(), os.path.join(config_folder_path, f"{fine_tuner}_lr{lr}_{loss}_batch_size_{batch_size}_step_size_{scheduler_step_size}_scheduler_gamma_{scheduler_gamma}.pth"))

def initiate(lrs: list[typing.Union[str, float]],
             combined: bool,
             train: bool,
             pretrained_path: str = None,
             debug: bool = False):
    print(f'Models: {vit_model_names}\n'
          f'Epochs num: {num_epochs}\n'
          f'Learning rates: {lrs}')

    datasets, num_fine_grain_classes, num_coarse_grain_classes = data_preprocessing.get_datasets(cwd=cwd)

    if combined:
        device = torch.device('cpu') if debug else (
            torch.device('mps' if torch.backends.mps.is_available() else
                         ("cuda" if torch.cuda.is_available() else 'cpu')))
        devices = [device]
        print(f'Using {device}')

        num_classes = num_fine_grain_classes + num_coarse_grain_classes

        if (not train) and (pretrained_path is not None):
            print(f'Loading pretrained model from {pretrained_path}')
            fine_tuners = [models.VITFineTuner.from_pretrained(vit_model_name=vit_model_name,
                                                               classes_num=num_classes,
                                                               pretrained_path=pretrained_path,
                                                               device=device)
                           for vit_model_name in vit_model_names]
        else:
            fine_tuners = [models.VITFineTuner(vit_model_name=vit_model_name,
                                               num_classes=num_classes)
                           for vit_model_name in vit_model_names]

        results_path = combined_results_path
    else:
        num_gpus = torch.cuda.device_count()

        if num_gpus < 2:
            raise ValueError("This setup requires at least 2 GPUs.")

        # Assign models to different GPUs
        devices = [torch.device("cuda:0"), torch.device("cuda:1")]

        fine_tuners = ([models.VITFineTuner(vit_model_name=vit_model_name,
                                            num_classes=num_fine_grain_classes)
                        for vit_model_name in vit_model_names] +
                       [models.VITFineTuner(vit_model_name=vit_model_name,
                                            num_classes=num_coarse_grain_classes)
                        for vit_model_name in vit_model_names])
        results_path = individual_results_path

    utils.create_directory(results_path)
    loaders = data_preprocessing.get_loaders(datasets=datasets,
                                             batch_size=batch_size)

    return fine_tuners, loaders, devices, num_fine_grain_classes, num_coarse_grain_classes


def run_combined_fine_tuning_pipeline(lrs: list[typing.Union[str, float]],
                                      loss: str = 'BCE',
                                      save_files: bool = True,
                                      debug: bool = utils.is_debug_mode()):
    fine_tuners, loaders, devices, num_fine_grain_classes, num_coarse_grain_classes = initiate(lrs=lrs,
                                                                                               combined=True,
                                                                                               train=True,
                                                                                               debug=debug)
    for fine_tuner in fine_tuners:
        with context_handlers.ClearSession():
            fine_tune_combined_model(lrs=lrs,
                                     fine_tuner=fine_tuner,
                                     device=devices[0],
                                     loaders=loaders,
                                     num_fine_grain_classes=num_fine_grain_classes,
                                     num_coarse_grain_classes=num_coarse_grain_classes,
                                     loss=loss,
                                     save_files=save_files,
                                     debug=debug)
            print('#' * 100)


def run_individual_fine_tuning_pipeline(lrs: list[typing.Union[str, float]],
                                        save_files: bool = True,
                                        debug: bool = utils.is_debug_mode()):
    fine_tuners, loaders, devices, num_fine_grain_classes, num_coarse_grain_classes = initiate(lrs=lrs,
                                                                                               combined=False,
                                                                                               train=True,
                                                                                               debug=debug)

    for fine_tuner in fine_tuners:
        print(f'Initiating {fine_tuner}')

        with context_handlers.ClearSession():
            fine_tune_individual_models(fine_tuners=fine_tuners,
                                        devices=devices,
                                        loaders=loaders,
                                        num_fine_grain_classes=num_fine_grain_classes,
                                        num_coarse_grain_classes=num_coarse_grain_classes,
                                        save_files=save_files)
            print('#' * 100)


def run_combined_testing_pipeline(lrs: list[typing.Union[str, float]],
                                  loss: str,
                                  pretrained_path: str = None,
                                  save_files: bool = True,
                                  debug: bool = utils.is_debug_mode()):
    fine_tuners, loaders, devices, num_fine_grain_classes, num_coarse_grain_classes = (
        initiate(lrs=lrs,
                 combined=True,
                 train=False,
                 pretrained_path=pretrained_path,
                 debug=debug))

    (test_fine_ground_truths, test_coarse_ground_truths, test_fine_predictions, test_coarse_predictions,
     test_fine_accuracy, test_coarse_accuracy) = test_combined_model(fine_tuner=fine_tuners[0],
                                                                     loaders=loaders,
                                                                     loss=loss,
                                                                     device=devices[0])

    if save_files:
        save_test_files(fine_tuners=fine_tuners[0],
                        combined=True,
                        lrs=lrs[0],
                        epoch=num_epochs,
                        test_fine_prediction=test_fine_predictions,
                        test_coarse_prediction=test_coarse_predictions)



def get_and_not_print_metrics_update(fine_predictions: np.array,
                          coarse_predictions: np.array,
                          loss: str,
                          true_fine_data: np.array = data_preprocessing.true_fine_data,
                          true_coarse_data: np.array = data_preprocessing.true_coarse_data,):
    test_fine_accuracy = accuracy_score(y_true=true_fine_data,
                                        y_pred=fine_predictions)
    test_coarse_accuracy = accuracy_score(y_true=true_coarse_data,
                                          y_pred=coarse_predictions)
    test_fine_f1 = f1_score(y_true=true_fine_data,
                            y_pred=fine_predictions,
                            labels=range(len(data_preprocessing.fine_grain_classes)),
                            average='macro')
    test_coarse_f1 = f1_score(y_true=true_coarse_data,
                              y_pred=coarse_predictions,
                              labels=range(len(data_preprocessing.coarse_grain_classes)),
                              average='macro')
    
    return test_fine_accuracy, test_fine_f1, test_coarse_accuracy, test_coarse_f1

def load_prediction_from_dropbox(
    config_folder_path, 
    loss, 
    lr, 
    beta, 
    batch_size, 
    scheduler_step_size, 
    scheduler_gamma
):
    """
    Load the model from the dropbox

    Args:
        config_folder_path: The path to the configuration folder.
        loss: The type of loss function used.
        lr: The learning rate used for training.
        beta: Beta value for LTN-based losses (if applicable).
        batch_size: The batch size used for training.
        scheduler_step_size: The step size for the learning rate scheduler.
        scheduler_gamma: The gamma value for the learning rate scheduler.
        test_fine_predictions: The fine-grained predictions on the test set.
        test_coarse_predictions: The coarse-grained predictions on the test set.
    """

    file_name = f"vit_b_16_lr{lr}_{loss}_beta_{beta}_batch_size_{batch_size}_step_size_{scheduler_step_size}_scheduler_gamma_{scheduler_gamma}"
    save_model_path = os.path.join(config_folder_path, file_name)

    # Load files
    file_based_to = save_model_path
    file_to_folder_loss = "BCE" if loss == "LTN_BCE" else "Softmarginal"
    file_to_folder_lr = "1e-4" if lr == 1e-4 else "3e-6"
    file_based_from = f"/EDCR/Results/combined/{file_to_folder_loss}/vit_b_16/lr_{file_to_folder_lr}/batch_size_{batch_size}/scheduler_gamma_{scheduler_gamma}/step_size_{scheduler_step_size}/beta_{beta}/"

    try:
        transferData.download_file(f"{file_based_from}{file_name}_fine_pred.npy", f"{file_based_to}_fine_pred.npy")
        transferData.download_file(f"{file_based_from}{file_name}_coarse_pred.npy", f"{file_based_to}_coarse_pred.npy")
    
        print(f"Predict fine and coarse and model is download to: {file_based_to} with file name: {file_name}")
    except dropbox.exceptions.ApiError as err:
        print(f"Error uploading file: {err}")

    return f"{file_based_to}_fine_pred.npy", f"{file_based_to}_coarse_pred.npy"

def test_and_save_LTN_combine_prediction(loss, lr, scheduler_step_size, scheduler_gamma, workbook_path=None):
    """
    Get prediction from dropbox file and send it to an excel file
    """

    base_path = os.getcwd()
    config_folder_name = f"vit_b_16_lr{lr}_{loss}_batch_size_{batch_size}_step_size_{scheduler_step_size}_scheduler_gamma_{scheduler_gamma}"
    config_folder_path = os.path.join(base_path, "model", config_folder_name)
    os.makedirs(config_folder_path, exist_ok=True)

    result = []

    for beta in betas:
        
        fine_pred_path, coarse_pred_path = load_prediction_from_dropbox(
            config_folder_path, 
            loss, 
            lr, 
            beta, 
            batch_size, 
            scheduler_step_size, 
            scheduler_gamma
        )

        # get result
        fine_predictions = np.load(fine_pred_path)
        coarse_predictions = np.load(coarse_pred_path)

        # Do EDCR on it to get the new fine and coarse prediction instead

        test_fine_accuracy, test_fine_f1, test_coarse_accuracy, test_coarse_f1 = get_and_not_print_metrics_update(fine_predictions, coarse_predictions, loss="")
        num_inconsistencies = data_preprocessing.get_num_inconsistencies(fine_predictions, coarse_predictions)
        percent_inconsistencies = num_inconsistencies / 1621 # TODO: change the number to len of dataset
        total_fine_coarse_accuracy = test_fine_accuracy + test_coarse_accuracy

        # add to list result

        result.append([
            round(beta, 2), 
            round(test_fine_accuracy * 100, 2), 
            round(test_fine_f1 * 100, 2), 
            round(test_coarse_accuracy * 100, 2), 
            round(test_coarse_f1 * 100, 2), 
            round(num_inconsistencies, 2), 
            round(percent_inconsistencies * 100, 2), 
            round(total_fine_coarse_accuracy * 100 / 2, 2)
        ])
        
    # the result will have the list contain what you need for. Help me create an excel file

    # Create Excel file and format it as shown in the screenshot
    if workbook_path:
        if os.path.exists(workbook_path):
            workbook = openpyxl.load_workbook(workbook_path)
        else:
            workbook = openpyxl.Workbook()
            workbook_name = os.path.basename(workbook_path)
            workbook.save(workbook_path)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Find the last non-empty row (starting from row 1)
    last_row = max(sheet.max_row, 1)  # Ensure at least one row
    while not any(sheet.cell(row=last_row, column=col).value for col in range(1, 8)):
        last_row -= 1
        if last_row == 0:  # Break if we reach row 0 (which doesn't exist)
            break

    # Set the base row three rows below the last non-empty row
    base_row = last_row + 4

    # Set the title as the folder_model_path
    sheet.merge_cells(f'A{base_row}:H{base_row}')  # Merge cells for title
    title_cell = sheet.cell(row=base_row, column=1)
    title_cell.value = os.path.basename(config_folder_name)
    title_cell.font = openpyxl.styles.Font(bold=True, size=14)


    # Write headers in bold
    headers = ['Beta', 'Fine-grain Prior Combined Accuracy', 'Fine-grain Prior Combined Average F1',
               'Coarse-grain Prior Combined Accuracy', 'Coarse-grain Prior Combined Average F1',
               'Total Prior Inconsistencies', 'Total Prior Inconsistencies (%)',
               'Fine and Coarse Accuracy']
    for i, header in enumerate(headers):
        cell = sheet.cell(row=base_row + 1, column=i + 1)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)

    # Sort the results based on beta in ascending order
    result.sort(key=lambda row: row[0])  # Sort based on the first element (beta)

    max_accuracy = 0
    min_inconsistencies = 0

    min_inconsistencies_row = None
    max_accuracy_row = None
    
    for row_num, row_data in enumerate(result, start=base_row + 2):
        for col_num, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
    
            # Find minimum Total Prior Inconsistencies (%) row (only once)
            if col_num == 7 and min_inconsistencies_row is None:
                min_inconsistencies_row = row_num
                min_inconsistencies = value
            elif col_num == 7 and value < min_inconsistencies:
                min_inconsistencies = value
                min_inconsistencies_row = row_num
    
            # Find maximum Fine and Coarse Accuracy row (only once)
            if col_num == 8 and max_accuracy_row is None:
                max_accuracy_row = row_num
                max_accuracy = value
            elif col_num == 8 and value > max_accuracy:
                max_accuracy = value
                max_accuracy_row = row_num

    # Highlight the found rows after the loop
    # Change text color of entire rows
    if min_inconsistencies_row is not None:
        for col_num in range(1, sheet.max_column + 1):  # Iterate through all columns
            cell = sheet.cell(row=min_inconsistencies_row, column=col_num)
            cell.font = openpyxl.styles.Font(color="00FF00")  # Green text color
    
    if max_accuracy_row is not None:
        for col_num in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=max_accuracy_row, column=col_num)
            cell.font = openpyxl.styles.Font(color="FF0000")  # Red text color

    # Save the Excel file
    workbook.save(workbook_path)

def generate_random_text(length=20):
    letters_and_digits = random.random()
    return str(letters_and_digits)


def test_dropbox():
    print("begin test dropbox")
    folder_name = "test_folder/"  # You can customize the folder name
    file_name = "test_file.txt"
    
    os.makedirs(folder_name, exist_ok=True)  # Create folder if it doesn't exist

    random_text = generate_random_text()
    
    with open(os.path.join(folder_name, file_name), "w") as f:
        f.write(random_text)  # Write random text to the file

    try:
        transferData.upload_file(os.path.join(folder_name, file_name), "/test/test.txt")
        print(random_text)
        print("File uploaded successfully!")
    except dropbox.exceptions.ApiError as err:
        print(f"Error uploading file: {err}")
    

if __name__ == '__main__':
    # test sending file to dropbox. If this code has error and stop training, change the tokenized to dropbox...
    # otherwise, the token work successfully
    test_dropbox()
    # here is the code for training
    # for beta in betas:
    #     fine_tuners, loaders, devices, num_fine_grain_classes, num_coarse_grain_classes = (
    #             initiate(lrs=lrs,
    #                         combined=True,
    #                         train=False,
    #                         pretrained_path= pretrained_path,
    #                         debug=False))
    #     (test_fine_ground_truths, test_coarse_ground_truths, test_fine_predictions, test_coarse_predictions,
    #             test_fine_accuracy, test_coarse_accuracy) = test_combined_model(fine_tuner=fine_tuners[0],
    #                                                                             loaders=loaders,
    #                                                                             loss="",
    #                                                                             device=devices[0])
    #     for fine_tuner in fine_tuners:
    #             with context_handlers.ClearSession():
    #                 fine_tune_combined_model(lrs=lrs,
    #                                             fine_tuner=fine_tuner,
    #                                             device=devices[0],
    #                                             loaders=loaders,
    #                                             num_fine_grain_classes=num_fine_grain_classes,
    #                                             num_coarse_grain_classes=num_coarse_grain_classes,
    #                                             loss=loss,
    #                                             ltn_num_epochs = ltn_num_epochs,
    #                                             beta=beta,
    #                                             save_files=True,
    #                                             debug=False)
    #                 print('#' * 100)
    for loss in ["LTN_BCE", "LTN_soft_marginal"]:
        for lr in [1e-04]:
            for scheduler_step_size in [1, 2]:
                for scheduler_gamma in [0.1, 0.3, 0.5, 0.8]:
                    test_and_save_LTN_combine_prediction(loss, lr, scheduler_step_size, scheduler_gamma, "/Users/khoavo2003/cs224/metacognitive_error_detection_and_correction_v2/LTN_Combine_Results.xlsx")
    






    

