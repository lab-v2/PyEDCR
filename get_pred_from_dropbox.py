import os
import torch.utils.data
import numpy as np
from sklearn.metrics import accuracy_score, f1_score
import pathlib
import typing
import random
import context_handlers
import models
import utils
import data_preprocessing
import re
import openpyxl  # Using openpyxl for Excel creation
from openpyxl.chart import (
    LineChart,
    Reference,
    Series,
)
import dropbox
from utils import TransferData


# token for Dropbox API. You can get the token from dropbox by go to dropbox:
# https://www.dropbox.com/developers/apps?_tk=pilot_lp&_ad=topbar4&_camp=myapps
# Notice that if your school account not work, use a different account
access_token = "Your token"
transferData = TransferData(access_token)

batch_size = 512
scheduler_gamma = 0.8
num_epochs = 5
ltn_num_epochs = 5
vit_model_names = [f'vit_{vit_model_name}' for vit_model_name in ['b_16']]
loss = "LTN_soft_marginal"
lrs=[3e-06]

files_path = '/content/drive/My Drive/' if utils.is_running_in_colab() else ''
combined_results_path = fr'{files_path}combined_results/'
individual_results_path = fr'{files_path}individual_results/'
cwd = pathlib.Path(__file__).parent.resolve()
scheduler_step_size = 1
betas = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]


def get_and_not_print_metrics_update(fine_predictions: np.array,
                          coarse_predictions: np.array,
                          loss: str,
                          true_fine_data: np.array = data_preprocessing.true_fine_data,
                          true_coarse_data: np.array = data_preprocessing.true_coarse_data,):
    test_fine_accuracy = accuracy_score(y_true=true_fine_data,
                                        y_pred=fine_predictions)
    test_coarse_accuracy = accuracy_score(y_true=true_coarse_data,
                                          y_pred=coarse_predictions)
    test_fine_f1 = f1_score(y_true=true_fine_data,
                            y_pred=fine_predictions,
                            labels=range(len(data_preprocessing.fine_grain_classes)),
                            average='macro')
    test_coarse_f1 = f1_score(y_true=true_coarse_data,
                              y_pred=coarse_predictions,
                              labels=range(len(data_preprocessing.coarse_grain_classes)),
                              average='macro')
    
    return test_fine_accuracy, test_fine_f1, test_coarse_accuracy, test_coarse_f1

def load_prediction_from_dropbox(
    config_folder_path, 
    loss, 
    lr, 
    beta, 
    batch_size, 
    scheduler_step_size, 
    scheduler_gamma
):
    """
    Load the model from the dropbox

    Args:
        config_folder_path: The path to the configuration folder.
        loss: The type of loss function used.
        lr: The learning rate used for training.
        beta: Beta value for LTN-based losses (if applicable).
        batch_size: The batch size used for training.
        scheduler_step_size: The step size for the learning rate scheduler.
        scheduler_gamma: The gamma value for the learning rate scheduler.
        test_fine_predictions: The fine-grained predictions on the test set.
        test_coarse_predictions: The coarse-grained predictions on the test set.
    """

    file_name = f"vit_b_16_lr{lr}_{loss}_beta_{beta}_batch_size_{batch_size}_step_size_{scheduler_step_size}_scheduler_gamma_{scheduler_gamma}"
    save_model_path = os.path.join(config_folder_path, file_name)

    # Load files
    file_based_to = save_model_path
    file_to_folder_loss = "BCE" if loss == "LTN_BCE" else "Softmarginal"
    file_to_folder_lr = "1e-4" if lr == 1e-4 else "3e-6"
    file_based_from = f"/EDCR/Results/combined/{file_to_folder_loss}/vit_b_16/lr_{file_to_folder_lr}/batch_size_{batch_size}/scheduler_gamma_{scheduler_gamma}/step_size_{scheduler_step_size}/beta_{beta}/"

    try:
        transferData.download_file(f"{file_based_from}{file_name}_fine_pred.npy", f"{file_based_to}_fine_pred.npy")
        transferData.download_file(f"{file_based_from}{file_name}_coarse_pred.npy", f"{file_based_to}_coarse_pred.npy")
    
        print(f"Predict fine and coarse and model is download to: {file_based_to} with file name: {file_name}")
    except dropbox.exceptions.ApiError as err:
        print(f"Error uploading file: {err}")

    return f"{file_based_to}_fine_pred.npy", f"{file_based_to}_coarse_pred.npy"

def test_and_save_LTN_combine_prediction(loss, lr, scheduler_step_size, scheduler_gamma, workbook_path=None):
    """
    Get prediction from dropbox file and send it to an excel file
    """

    base_path = os.getcwd()
    config_folder_name = f"vit_b_16_lr{lr}_{loss}_batch_size_{batch_size}_step_size_{scheduler_step_size}_scheduler_gamma_{scheduler_gamma}"
    config_folder_path = os.path.join(base_path, "model", config_folder_name)
    os.makedirs(config_folder_path, exist_ok=True)

    result = []

    for beta in betas:
        
        fine_pred_path, coarse_pred_path = load_prediction_from_dropbox(
            config_folder_path, 
            loss, 
            lr, 
            beta, 
            batch_size, 
            scheduler_step_size, 
            scheduler_gamma
        )

        # get result
        fine_predictions = np.load(fine_pred_path)
        coarse_predictions = np.load(coarse_pred_path)

        # Do EDCR on it to get the new fine and coarse prediction instead

        test_fine_accuracy, test_fine_f1, test_coarse_accuracy, test_coarse_f1 = get_and_not_print_metrics_update(fine_predictions, coarse_predictions, loss="")
        num_inconsistencies = data_preprocessing.get_num_inconsistencies(fine_predictions, coarse_predictions)
        percent_inconsistencies = num_inconsistencies / 1621 # TODO: change the number to len of dataset
        total_fine_coarse_accuracy = test_fine_accuracy + test_coarse_accuracy

        # add to list result

        result.append([
            round(beta, 2), 
            round(test_fine_accuracy * 100, 2), 
            round(test_fine_f1 * 100, 2), 
            round(test_coarse_accuracy * 100, 2), 
            round(test_coarse_f1 * 100, 2), 
            round(num_inconsistencies, 2), 
            round(percent_inconsistencies * 100, 2), 
            round(total_fine_coarse_accuracy * 100 / 2, 2)
        ])
        
    # the result will have the list contain what you need for. Help me create an excel file

    # Create Excel file and format it as shown in the screenshot
    if workbook_path:
        if os.path.exists(workbook_path):
            workbook = openpyxl.load_workbook(workbook_path)
        else:
            workbook = openpyxl.Workbook()
            workbook_name = os.path.basename(workbook_path)
            workbook.save(workbook_path)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Find the last non-empty row (starting from row 1)
    last_row = max(sheet.max_row, 1)  # Ensure at least one row
    while not any(sheet.cell(row=last_row, column=col).value for col in range(1, 8)):
        last_row -= 1
        if last_row == 0:  # Break if we reach row 0 (which doesn't exist)
            break

    # Set the base row three rows below the last non-empty row
    base_row = last_row + 4

    # Set the title as the folder_model_path
    sheet.merge_cells(f'A{base_row}:H{base_row}')  # Merge cells for title
    title_cell = sheet.cell(row=base_row, column=1)
    title_cell.value = os.path.basename(config_folder_name)
    title_cell.font = openpyxl.styles.Font(bold=True, size=14)


    # Write headers in bold
    headers = ['Beta', 'Fine-grain Prior Combined Accuracy', 'Fine-grain Prior Combined Average F1',
               'Coarse-grain Prior Combined Accuracy', 'Coarse-grain Prior Combined Average F1',
               'Total Prior Inconsistencies', 'Total Prior Inconsistencies (%)',
               'Fine and Coarse Accuracy']
    for i, header in enumerate(headers):
        cell = sheet.cell(row=base_row + 1, column=i + 1)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)

    # Sort the results based on beta in ascending order
    result.sort(key=lambda row: row[0])  # Sort based on the first element (beta)

    max_accuracy = 0
    min_inconsistencies = 0

    min_inconsistencies_row = None
    max_accuracy_row = None
    
    for row_num, row_data in enumerate(result, start=base_row + 2):
        for col_num, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
    
            # Find minimum Total Prior Inconsistencies (%) row (only once)
            if col_num == 7 and min_inconsistencies_row is None:
                min_inconsistencies_row = row_num
                min_inconsistencies = value
            elif col_num == 7 and value < min_inconsistencies:
                min_inconsistencies = value
                min_inconsistencies_row = row_num
    
            # Find maximum Fine and Coarse Accuracy row (only once)
            if col_num == 8 and max_accuracy_row is None:
                max_accuracy_row = row_num
                max_accuracy = value
            elif col_num == 8 and value > max_accuracy:
                max_accuracy = value
                max_accuracy_row = row_num

    # Highlight the found rows after the loop
    # Change text color of entire rows
    if min_inconsistencies_row is not None:
        for col_num in range(1, sheet.max_column + 1):  # Iterate through all columns
            cell = sheet.cell(row=min_inconsistencies_row, column=col_num)
            cell.font = openpyxl.styles.Font(color="00FF00")  # Green text color
    
    if max_accuracy_row is not None:
        for col_num in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=max_accuracy_row, column=col_num)
            cell.font = openpyxl.styles.Font(color="FF0000")  # Red text color

    # Save the Excel file
    workbook.save(workbook_path)